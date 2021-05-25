import openpyxl as xl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from fuzzywuzzy import fuzz
from copy import copy
import numpy as np
import re
PAGE_LINE = 64
data = xl.open("/home/vuong/PycharmProjects/hand-predict/goyou/新石垣航空基地.xlsx")
worksheet = data["工事内訳_1"]
TEXT = ["名　　　　　　　　　　称", "数　　量", "単位", "金　　　　　額", "備　　　考"]
numbers_of_page = (worksheet.max_row + 1) // PAGE_LINE
redFill = PatternFill(start_color='FFFF0000',
                      end_color='FFFF0000',
                      fill_type='solid')
def compare_sheet(sheet, title, str, page):
    if (sheet.cell(page * PAGE_LINE + 2, 2).value != title):
        print(title)
        print(" ")
        sheet.cell(page * PAGE_LINE + 2, 2).fill = redFill
    if (sheet.cell(page * PAGE_LINE + 1, 2).value != str):
        sheet.cell(page * PAGE_LINE + 1, 2).fill = redFill
    if (sheet.cell(page * PAGE_LINE + 1, 2).value != str):
        sheet.cell(page * PAGE_LINE + 1, 2).fill = redFill
def compare_sheet_fuzzy(sheet, title, title_2, str, page):
    print(title)
    print(title_2.replace("  ", ""))
    print(page + 46)
    print(" ")
    if fuzz.ratio(title_2.replace("  ", ""), title) < 95:
        sheet.cell(page * PAGE_LINE + 2, 2).fill = redFill
    if (sheet.cell(page * PAGE_LINE + 1, 2).value != str):
        sheet.cell(page * PAGE_LINE + 1, 2).fill = redFill
    if (sheet.cell(page * PAGE_LINE + 1, 2).value != str):
        sheet.cell(page * PAGE_LINE + 1, 2).fill = redFill


titles = []
def write_page_sheet1(page, title, sheet):
    sheet.cell(page * PAGE_LINE + 2, 2).value = title
    sheet.cell(page * PAGE_LINE + 1, 2).value = "直接工事費科目別内訳"
    sheet.cell(page * PAGE_LINE + 2, 6).value = "PAGE " + (page + 4).__str__()
    titles.append(title)
    compare_sheet(sheet, title, "直接工事費科目別内訳", page)

    for j, title in enumerate(TEXT):
        sheet.cell(page * PAGE_LINE + 3, j + 2).value = title
        for i in range(3, 65):
            val = sheet.cell(page * PAGE_LINE + i + 1, j + 2).value
            sheet.cell(page * PAGE_LINE + i + 1, j + 2).value = val


count = 2
for row in range(68, 127):
    val = worksheet.cell(row, 2).value
    if (val is not None) and (val != "計"):
        write_page_sheet1(count, val, worksheet)
        count = count + 1

######################SHEET2###############################
count = 0
worksheet = data["中科目別内訳_1"]
title_sheet3 = []
val_sheet3 = []

def write_page_sheet2(page, title, sheet):
    while True:
        if (fuzz.ratio(title, sheet.cell(page * PAGE_LINE + 2, 2).value) < 85) :
            break

        compare_sheet(sheet, title, "直接工事費中科目別内訳", page)
        sheet.cell(page * PAGE_LINE + 2, 2).value = title
        sheet.cell(page * PAGE_LINE + 1, 2).value = "直接工事費中科目別内訳"
        sheet.cell(page * PAGE_LINE + 2, 7).value = "PAGE " + (page + 22).__str__()
        for j, tit in enumerate(TEXT):
            sheet.cell(page * PAGE_LINE + 3, j + 2).value = tit
            for i in range(3, 64):
                val = sheet.cell(page * PAGE_LINE + i + 1, j + 2).value
                val2 = sheet.cell(page * PAGE_LINE + i + 1, j + 3).value
                if val2 is None:
                    val2 = ''
                else:
                    val2 = "  " + val2
                if (j == 0) and (val is not None) and (val != "計") and (re.findall("([ぁ-んァ-ン\w])", val) != []):
                    title_sheet3.append(title + "  " + val + val2)
                    val_sheet3.append(sheet.cell(page * PAGE_LINE + i + 2, j + 6).value)
        page += 1

    return page

for title in titles:
    count = write_page_sheet2(count, title, worksheet)

#####################SHEET3#############################################
worksheet = data["別紙明細_1"]

value_sheet3_5 = []
count = 0

title_sheet_3_5 = []
def write_page_sheet3(page, title, sheet, j):
    while True:
        a = title[1] + "".join(re.findall("([ぁ-んァ-ン\w])", title ))
        b = sheet.cell(page * PAGE_LINE + 2, 2).value[1] + "".join(re.findall("([ぁ-んァ-ン\w])", sheet.cell(page * PAGE_LINE + 2, 2).value ))
        for i in range(4,64):
            if sheet.cell(page * PAGE_LINE + i, 7).value is not None :
                v = sheet.cell(page * PAGE_LINE + i, 7).value
            if sheet.cell(page * PAGE_LINE + i, 2).value is not None:
                check = sheet.cell(page * PAGE_LINE + i, 2).value


        compare_sheet_fuzzy(sheet, title.replace("  ",""), sheet.cell(page * PAGE_LINE + 2, 2).value, "直接工事費細目別内訳", page)
        for (i, t) in enumerate(title.split("  ")):
            if i == 2:
                sheet.cell(page * PAGE_LINE + 2, 7).value = t
                continue
            sheet.cell(page * PAGE_LINE + 2, 2 + i * 2).value = t
        sheet.cell(page * PAGE_LINE + 1, 2).value = "直接工事費細目別内訳"
        sheet.cell(page * PAGE_LINE + 2, 8).value = "PAGE " + (page + 46).__str__()
        for i in range(3, 65):
            val = sheet.cell(page * PAGE_LINE + i + 1, 7).value
            if (val != "金　　　　額") and (val is not None):
                value_sheet3_5.append(val)
                title_sheet_3_5.append(title)
        page += 1
        if (title == "(6)ハイドラント(1)・"):
            print(2)
        if (fuzz.ratio(a,b) > 85) and (v == val_sheet3[j]) :
            break

    return page

def get_val(sheet, page, ):
    ans = []
    for i in range(page* PAGE_LINE, sheet.max_row + 1):
        if (i%64 != 5):
            continue
        val = sheet.cell(i, 7).value
        ans.append(val)
    return ans
def write_page_sheet3_5(page, pred_val, sheet ):
    for i,val in enumerate(value_sheet3_5):
        t = sheet.cell(page * PAGE_LINE + 2, 2).value
        if val != pred_val or (fuzz.ratio(title_sheet_3_5[i].replace("  ", ""), t) < 85):
            continue
        compare_sheet_fuzzy(sheet, title_sheet_3_5[i].replace("  ", ""), t, "直接工事費別紙明細", page)
        for (i, t) in enumerate(title_sheet_3_5[i].split("  ")):
            if i == 2:
                sheet.cell(page * PAGE_LINE + 2, 7).value = t
                continue
            sheet.cell(page * PAGE_LINE + 2, 2 + i * 2).value = t
        sheet.cell(page * PAGE_LINE + 1, 2).value = "直接工事費別紙明細"
        sheet.cell(page * PAGE_LINE + 2, 8).value = "PAGE " + (page + 46).__str__()
        break


for j,title in enumerate(title_sheet3):
    count = write_page_sheet3(count, title, worksheet, j)
pred_val_35 = get_val(sheet= worksheet, page= count)
pred_val_3 = get_val(sheet=worksheet, page =0)
for (i, pred_title) in enumerate(pred_val_35):
    write_page_sheet3_5(i + count, pred_title, worksheet)

################################SHEET3########################################
data.save("test.xls")

